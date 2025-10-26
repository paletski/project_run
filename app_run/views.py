from rest_framework.decorators import api_view
from rest_framework import viewsets

from django.conf import settings

from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem
from .serializers import RunSerializer, UserSerializer, AthleteInfoSerializer, \
    ChallengeSerializer, PositionSerializer, CollectibleItemSerializer, \
    UserSerializerCollItems
from .serializers import ChallengeSerializer
from django.contrib.auth.models import User
from rest_framework.filters import SearchFilter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from rest_framework.pagination import PageNumberPagination
from geopy.distance import geodesic
from django.db.models import Sum, Min, Max, Count, Q
from openpyxl import load_workbook
import copy
# Create your views here.

@api_view(['GET'])
def company_description_view(request):
    return Response({'company_name': settings.COMPANY_NAME,
                    'slogan': settings.SLOGAN,
                    'contacts': settings.CONTACTS}
                    )

class RunPagination(PageNumberPagination):
    page_size_query_param = 'size'


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    pagination_class = RunPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete'] # фильтруем
    ordering_fields = ['created_at']         # сортируем

class UserPagination(PageNumberPagination):
    page_size_query_param = 'size'


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    #queryset = User.objects.all()
    queryset = User.objects.annotate(runs_finished=Count('runs', filter=Q(
        runs__status='finished'))).distinct()
    #print(f'queryset = {queryset}')

    serializer_class = UserSerializer
    pagination_class = RunPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    search_fields = ['first_name', 'last_name',]
    ordering_fields = ['date_joined']

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            return UserSerializerCollItems
        return super().get_serializer_class()

    def get_queryset(self):
        qs = self.queryset
        #print(f'queryset_inn = {qs}')
        type = self.request.query_params.get('type', None)
        if type:
            if type == 'coach':
                qs = qs.filter(is_staff=True, is_superuser=False)
            elif type == 'athlete':
                qs = qs.filter(is_staff=False, is_superuser=False)
            else:
                qs = qs.filter(is_superuser=False)
        return qs.filter(is_superuser=False)



class RunStartViewSet(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        #print(f'view start {run}')
        if run.status == 'init':
            run.status = 'in_progress'
            run.save()
            data = {'message': f'POST запрос обработан 3'}
            return Response(data, status=status.HTTP_200_OK)
        else:
            data = {'message': f'POST запрос не обработан 2'}
            return Response(data, status=status.HTTP_400_BAD_REQUEST)


class RunStopViewSet(APIView):
    # получаем список всех загруженных коллектитемсов
    collectible_items_list = CollectibleItem.objects.all()
    #print(collectible_items_list)
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        #print(f'view stop {run}')
        if run.status == 'in_progress':
            run.status = 'finished'
            # сюда добавим вычисления расстояния забега
            positions_list = Position.objects.filter(run=run).order_by('date_time')
            #print(positions_list)

            # вычислим длительность забега
            min_time = Position.objects.filter(run=run).aggregate(Min('date_time'))
            max_time = Position.objects.filter(run=run).aggregate(Max('date_time'))
            min_dt = min_time['date_time__min']
            max_dt = max_time['date_time__max']
            # Как посчитать в базе? все равно мин/макс ищет?
            if not max_dt or not min_dt:
                run_time = None
            else:
                run_time = int((max_dt - min_dt).total_seconds())
            # получить id юзера, делающего забег
            user_id = run.athlete.id
            #print(f'user_id = {user_id}')

            # select-related ? хотя вроде зачем, мы же тянем все из позиций
            probeg = 0
            pos_cnt = positions_list.count()
            for i in range(pos_cnt):
                if i == 0:
                    probeg = 0
                    pos_new = (positions_list[0].latitude,
                               positions_list[0].longitude)
                    # нет ли рядом предметов CollectableItem?
                    if self.get_dist(pos_new)[0]:
                        coll_item = CollectibleItem.objects.get(id=self.get_dist(pos_new)[1])
                        coll_item.collitems.add(user_id)
                else:
                    pos_old = (positions_list[i-1].latitude, positions_list[i-1].longitude)
                    pos_new = (positions_list[i].latitude, positions_list[i].longitude)
                    probeg += geodesic(pos_old, pos_new).km

                    # нет ли рядом предметов CollectableItem?
                    if self.get_dist(pos_new)[0]:
                        coll_item = CollectibleItem.objects.get(id=self.get_dist(pos_new)[1])
                        coll_item.collitems.add(user_id)

            run.distance = probeg
            run.run_time_seconds = run_time
            run.save()

            # челленджи
            #
            cnt = Run.objects.filter(athlete=run.athlete, status='finished').count()
            # 'Сделай 10 Забегов!'
            if cnt == 10:  #даже если 11 забегов то с равенством ТЗ выполняется
                run_challenge = Challenge.objects.create(
                    full_name='Сделай 10 Забегов!', athlete=run.athlete)


            # "Пробеги 50 километров!"
            # TODO надо переделать, сначала проверяем есть ли выполненный
            #  челлендж, если есть - идем мимо, нет - считаем в базе
            distance_50 = Run.objects.filter(athlete=run.athlete,
                                             status='finished'
                                             ).aggregate(Sum('distance'))
            #print(f'distance_50 = {distance_50}')
            if distance_50.get('distance__sum') >= 50:
                if not Challenge.objects.filter(athlete=run.athlete,
                                                full_name='Пробеги 50 километров!').exists():
                    Challenge.objects.create(
                        full_name='Пробеги 50 километров!', athlete=run.athlete)


            data = {'message': f'POST запрос обработан 32'}
            return Response(data, status=status.HTTP_200_OK)
        else:
            data = {'message': f'POST запрос не обработан 22'}
            return Response(data, status=status.HTTP_400_BAD_REQUEST)

    def get_dist(self, position):
        col_item_cnt = self.collectible_items_list.count()
        # проходим по всем предметам, считаем расстояние от
        # текущей точки (position) до предмета, если менее 100 м добавить в
        # таблицу
        for j in range(col_item_cnt):
            pos_item = (self.collectible_items_list[j].latitude,
                        self.collectible_items_list[j].longitude)
            dist = geodesic(pos_item, position).m
            if dist < 100:
                return  [True, self.collectible_items_list[j].id]
        return [False, 0]


class AthleteInfoViewSet(APIView):
    queryset = AthleteInfo.objects.all()
    serializer_class = AthleteInfoSerializer
    def get(self, request, user_id):
        #True создан, False найден
        if not User.objects.filter(id=user_id).exists():
            message = {'message': 'пользователь не найден'}
            return Response(message, status=status.HTTP_404_NOT_FOUND)
        # юзер есть в User
        # значит достаем данные по нему из AthleteInfo или создаем запись в AthleteInfo
        athlete, created = AthleteInfo.objects.get_or_create(user_id_id=user_id)
        serializer = AthleteInfoSerializer(athlete)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, user_id):
        # True создан,  False обновлен
        if not User.objects.filter(id=user_id).exists():
            message = {'message': 'пользователь не найден'}
            return Response(message, status=status.HTTP_404_NOT_FOUND)

        athlete, created = AthleteInfo.objects.update_or_create(user_id_id=user_id)
        data = request.data
        if 'goals' in data:
            #print('goals in data')
            athlete.goals = data['goals']
        if 'weight' in data:
            try:
                weight = int(data['weight'])  # пока без try
                if not (0 < weight < 900):
                    message = {'msg': 'вес не в диапазоне 1-900'}
                    return Response(message, status=status.HTTP_400_BAD_REQUEST)
            # проскочили проверку, в диапазоне и число (помним про try)
            except ValueError:
                message = {'msg': 'вес должен быть числом в диапазоне 1-899'}
                return Response(message, status=status.HTTP_400_BAD_REQUEST)
            athlete.weight = weight
            athlete.save()
            serializer = AthleteInfoSerializer(athlete)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

class ChallengeViewSet(APIView):
    queryset = Challenge.objects.all()
    #print(queryset)
    serializer_class = ChallengeSerializer
    #def get(self, request, athlete = None):
    def get(self, request):
        athlete = request.query_params.get('athlete')
        #print(f'athlete = {athlete}')
        challenges = Challenge.objects.filter(athlete=athlete) if athlete else Challenge.objects.all()
        serializer = ChallengeSerializer(challenges,  many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        if 'run' in self.request.query_params:
            return self.queryset.filter(run_id=self.request.query_params['run'])
        else:
            return Position.objects.all()

    def create(self, request, *args, **kwargs):
        if 'run' in request.data:
            if Run.objects.filter(id=request.data['run'], status='in_progress').exists():
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response({'message': 'run не найден'},
                                status=status.HTTP_400_BAD_REQUEST)


class CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer



class CollectibleItemFileLoad(APIView):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer
    def post(self, request):
        wr_book = load_workbook(request.FILES['file'])
        wr_list = wr_book.active # берет первый лист - оно нам и надо
        # упрощенная проверка, по первому полю, если пусто или нет, то считаем
        # что строки кончились
        str_cnt = 2
        data_str = []
        data_val = {}
        err_str = []
        while True:
            values = [
                wr_list.cell(row=str_cnt, column=1).value, # name
                wr_list.cell(row=str_cnt, column=2).value, # uid
                wr_list.cell(row=str_cnt, column=3).value, # value
                wr_list.cell(row=str_cnt, column=4).value, # latitude
                wr_list.cell(row=str_cnt, column=5).value, # longitude
                wr_list.cell(row=str_cnt, column=6).value  # picture -  url в
            # экселе
            ]
            if values[0] is None or str(values[0]).strip() == "":
                break
            data_str.extend(values)
            # аналогично словарь:
            data_val = ({'name':      values[0],
                         'uid':       values[1],
                         'value':     values[2],
                         'latitude':  values[3],
                         'longitude': values[4],
                         'picture':   values[5]})

            serializer = CollectibleItemSerializer(data=data_val)
            if serializer.is_valid():
                if not CollectibleItem.objects.filter(uid= values[0]).exists():
                    serializer.save()
            else:
                err_str.append(copy.deepcopy(data_str))
            str_cnt += 1
            data_str.clear()
        return Response(err_str, status=status.HTTP_200_OK)



